import sys
import re
import asyncio
import aiohttp
import os
import time
import winsound
# import pygame
from itertools import zip_longest
from PySide6.QtWidgets import (
    QApplication,
    QWidget,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QHBoxLayout,
)
from PySide6.QtCore import QObject, QRunnable, QThreadPool, Signal, Qt
from openpyxl import Workbook
import openpyxl

# 添加了失败，成功的导出
# 添加了重试策略，加入了一行取下面状态的功能
# Pyinstaller -F -w 链接查订单-手动.py
#
# pygame.mixer.init()
# sound = pygame.mixer.Sound('apple-Query_model/测试/1.wav')


class AsyncRequest(QObject):
    finished = Signal()
    result = Signal(int, dict)

    def __init__(self, row, url):
        super().__init__()
        self.row = row
        self.url = url

    async def fetch_data(self, max_retries=100):
        for attempt in range(max_retries):
            try:
                async with aiohttp.ClientSession() as session:
                    response = await session.get(self.url, timeout=5)
                    response.raise_for_status()
                    response_text = await response.text()

                    # 正则表达式匹配逻辑
                    if '"Track shipment” link' in response_text:
                        Track_shipment = "Track shipment"
                    else:
                        Track_shipment = ""

                    modo = r'"productName"\s*:\s*"([^"]+)"'
                    time = r'"orderPlacedDate"\s*:\s*"([^"]+)"'
                    dd = r'/shop/order/detail/\d+/([^"/]+)'
                    status = r'"deliveryDate"\s*:\s*"([^"]+)"'
                    pattern = r'"currentStatus"\s*:\s*"([^"]+)"'
                    Order_number = r"trackingURLMap\":\s*\{\"(.*?)\"\:\s*\"(.*?)\""
                    # 仓库
                    first_names = re.findall(r'"firstName":"([^"]+)"', response_text)
                    last_names = re.findall(r'"lastName":"([^"]+)"', response_text)
                    name_pairs = [
                        f"{name1.strip()} {name2.strip()}"
                        for name1, name2 in zip_longest(
                            first_names, last_names, fillvalue=""
                        )
                    ]
                    unique_name_pairs = list(set(name_pairs))
                    sorted_unique_name_pairs = sorted(
                        unique_name_pairs,
                        key=lambda pair: (len(pair.split()[0]), pair.lower()),
                    )
                    first_variable = (
                        sorted_unique_name_pairs[0] if sorted_unique_name_pairs else ""
                    )
                    second_variable = (
                        sorted_unique_name_pairs[1]
                        if len(sorted_unique_name_pairs) > 1
                        else ""
                    )

                    MODO = re.search(modo, response_text)
                    TIME = re.search(time, response_text)
                    DD = re.search(dd, response_text)
                    STATUS = re.search(status, response_text)
                    The_status_of_the_goods = re.search(pattern, response_text)

                    Order_number = re.search(Order_number, response_text)

                    if MODO and TIME and DD and STATUS:
                        result_data = {
                            "productName": MODO.group(1) if MODO else "",
                            "orderPlacedDate": TIME.group(1) if TIME else "",
                            "orderNumber": DD.group(1) if DD else "",
                            "status": STATUS.group(1) if STATUS else "",
                            "Track_shipment": Track_shipment,
                            "The_status_of_the_goods": (
                                The_status_of_the_goods.group(1)
                                if The_status_of_the_goods
                                else ""
                            ),
                            "Order_number": (
                                Order_number.group(1) if Order_number else ""
                            ),
                            "Delivers": first_variable,
                            "Bills_to": second_variable,
                        }
                        # print(result_data)
                        self.result.emit(self.row, result_data)
                        return  # 成功获取数据后退出函数

                    raise ValueError("Data not found in response")

            except Exception as e:
                if attempt < max_retries:
                    print(attempt)
                    # await asyncio.sleep(2)
                    pass
                else:
                    self.result.emit(self.row, error_message)  # 发射错误信息
                error_message = {"error": f"Error: {e}"}
            finally:
                self.finished.emit()


class AsyncRequestWorker(QRunnable):
    def __init__(self, async_request):
        super().__init__()
        self.async_request = async_request

    def run(self):
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        loop.run_until_complete(self.async_request.fetch_data())


class InterFace(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.threadpool = QThreadPool()

    def initUI(self):
        self.setWindowTitle("Async HTTP Requests with PyQt6")
        self.setGeometry(100, 100, 1800, 600)

        h_layout_buttons = QHBoxLayout()
        self.button_import1 = QPushButton("导入", self)
        self.button_import2 = QPushButton("清理", self)
        self.button_import3 = QPushButton("查询", self)
        self.button_import4 = QPushButton("导出成功", self)
        self.button_import5 = QPushButton("导出失败", self)
        h_layout_buttons.addWidget(self.button_import1)
        h_layout_buttons.addWidget(self.button_import2)
        h_layout_buttons.addWidget(self.button_import3)
        h_layout_buttons.addWidget(self.button_import4)
        h_layout_buttons.addWidget(self.button_import5)
        h_layout_buttons.setAlignment(Qt.AlignmentFlag.AlignLeft)

        self.table_widget = QTableWidget(0, 10, self)
        self.table_widget.setHorizontalHeaderLabels(
            [
                "url",
                "modo",
                "order number",
                "order time",
                "status",
                "The_status_of_the_goods",
                "Order_number",
                "Track shipment",
                "Delivers",
                "Bills to",
            ]
        )
        self.table_widget.setShowGrid(True)
        self.table_widget.setColumnWidth(0, 500)
        self.table_widget.setColumnWidth(1, 200)
        self.table_widget.setColumnWidth(2, 100)
        self.table_widget.setColumnWidth(3, 100)
        self.table_widget.setColumnWidth(4, 150)
        self.table_widget.setColumnWidth(5, 200)
        self.table_widget.setColumnWidth(6, 150)
        self.table_widget.setColumnWidth(7, 200)
        self.table_widget.setColumnWidth(8, 70)
        self.table_widget.setColumnWidth(9, 70)

        v_layout = QVBoxLayout()
        v_layout.addLayout(h_layout_buttons)
        v_layout.addWidget(self.table_widget)
        self.setLayout(v_layout)

        self.button_import1.clicked.connect(self.Get_link)
        self.button_import2.clicked.connect(self.clear_table)
        self.button_import3.clicked.connect(self.on_query_clicked)
        self.button_import4.clicked.connect(self.export_data_to_xlsx_TRUE)
        self.button_import5.clicked.connect(self.export_data_to_xlsx_false)

    def on_query_clicked(self):
        for row in range(self.table_widget.rowCount()):
            for col in range(1, 10):
                self.table_widget.setItem(row, col, QTableWidgetItem(""))  # 清空列数据

        for row in range(self.table_widget.rowCount()):
            url_item = self.table_widget.item(row, 0)
            if url_item and url_item.text():
                async_request = AsyncRequest(row, url_item.text())
                async_request.result.connect(self.update_table)
                worker = AsyncRequestWorker(async_request)
                self.threadpool.start(worker)

    def update_table(self, row, result):
        if "error" in result:
            self.table_widget.setItem(row, 1, QTableWidgetItem(""))
            self.table_widget.setItem(row, 2, QTableWidgetItem(""))
            self.table_widget.setItem(row, 3, QTableWidgetItem(""))
            self.table_widget.setItem(row, 4, QTableWidgetItem(""))
            self.table_widget.setItem(row, 5, QTableWidgetItem(""))
            self.table_widget.setItem(row, 6, QTableWidgetItem(""))
            self.table_widget.setItem(row, 7, QTableWidgetItem(""))
            self.table_widget.setItem(row, 8, QTableWidgetItem(""))
            self.table_widget.setItem(row, 9, QTableWidgetItem(""))
        else:
            self.table_widget.setItem(
                row, 1, QTableWidgetItem(str(result["productName"]))
            )
            self.table_widget.setItem(
                row, 2, QTableWidgetItem(str(result["orderNumber"]))
            )
            self.table_widget.setItem(
                row, 3, QTableWidgetItem(str(result["orderPlacedDate"]))
            )
            self.table_widget.setItem(row, 4, QTableWidgetItem(str(result["status"])))
            self.table_widget.setItem(
                row, 5, QTableWidgetItem(str(result["The_status_of_the_goods"]))
            )
            self.table_widget.setItem(
                row, 6, QTableWidgetItem(str(result["Order_number"]))
            )
            self.table_widget.setItem(
                row, 7, QTableWidgetItem(str(result["Track_shipment"]))
            )
            self.table_widget.setItem(row, 8, QTableWidgetItem(str(result["Delivers"])))
            self.table_widget.setItem(row, 9, QTableWidgetItem(str(result["Bills_to"])))
            # if result['The_status_of_the_goods'] == "DELIVERED":
            #     winsound.Beep(800, 500)
            # !!!
            # if result['The_status_of_the_goods'] == "DELIVERED":
            #         sound.play()

    def Get_link(self):
        clipboard = QApplication.clipboard().text()
        if clipboard:
            items = clipboard.split()
            self.table_widget.setRowCount(len(items))
            for i, item in enumerate(items):
                self.table_widget.setItem(i, 0, QTableWidgetItem(item))

    def clear_table(self):
        self.table_widget.clearContents()
        self.table_widget.setRowCount(0)

    def export_data_to_xlsx_TRUE(self):

        current_time = time.strftime("%Y%m%d_%H%M%S")
        file_name = f"成功_{current_time}.xlsx"
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        file_path = os.path.join(desktop_path, file_name)

        wb = Workbook()
        ws = wb.active

        headers = [
            "url",
            "modo",
            "order number",
            "order time",
            "status",
            "The status of the goods",
            "Order_number",
            "Track shipment",
            "Delivers",
            "Bills to",
        ]
        ws.append(headers)

        for row in range(self.table_widget.rowCount()):
            empty_columns = all(
                self.table_widget.item(row, col) is None
                or self.table_widget.item(row, col).text() == ""
                for col in [1, 2, 3, 4, 5, 6, 7, 8, 9]
            )
            if not empty_columns:
                row_data = [
                    (
                        self.table_widget.item(row, col).text()
                        if self.table_widget.item(row, col)
                        else ""
                    )
                    for col in range(self.table_widget.columnCount())
                ]
                ws.append(row_data)

        for col in ws.columns:
            max_length = max(
                len(str(cell.value)) for cell in col if cell.value is not None
            )
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col[0].column)
            ].width = max_length

        # 保存文件
        wb.save(file_path)

    def export_data_to_xlsx_false(self):
        current_time = time.strftime("%Y%m%d_%H%M%S")
        file_name = f"失败_{current_time}.xlsx"
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        file_path = os.path.join(desktop_path, file_name)

        wb = Workbook()
        ws = wb.active

        # 添加列标题
        headers = [
            "url",
            "modo",
            "order number",
            "order time",
            "status",
            "The status of the goods",
            "Track shipment",
        ]
        ws.append(headers)

        # 遍历表中的每一行
        for row in range(self.table_widget.rowCount()):
            # 检查是否包含 'Cancelled' 或者至少有一个空列
            status_item = self.table_widget.item(row, 4)  # 假设第5列是状态列
            empty_columns = all(
                self.table_widget.item(row, col) is None
                or self.table_widget.item(row, col).text() == ""
                for col in range(1, 7)  # 假设第2列到第7列是需要检查的列
            )

            if empty_columns or (status_item and status_item.text() == "Cancelled"):
                row_data = [
                    (
                        self.table_widget.item(row, col).text()
                        if self.table_widget.item(row, col)
                        else ""
                    )
                    for col in range(self.table_widget.columnCount())
                ]
                ws.append(row_data)

        # 自动调整列宽
        for col in ws.columns:
            max_length = max(
                len(str(cell.value)) for cell in col if cell.value is not None
            )
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col[0].column)
            ].width = max_length

        # 保存工作簿
        wb.save(file_path)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = InterFace()
    window.show()
    sys.exit(app.exec())
