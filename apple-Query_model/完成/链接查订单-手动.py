import sys
import re
import asyncio
import aiohttp
import os
import time
from PyQt6.QtWidgets import QApplication,QWidget,QPushButton,QTableWidget,QTableWidgetItem,QVBoxLayout,QHBoxLayout
from PyQt6.QtCore import QObject, QRunnable, QThreadPool, pyqtSignal, Qt
from openpyxl import Workbook
import openpyxl


# 添加了失败，成功的导出
# 添加了重试策略，加入了一行取下面状态的功能
# Pyinstaller -F -w 链接查订单-手动.py
#


class AsyncRequest(QObject):
    finished = pyqtSignal()
    result = pyqtSignal(int, dict)

    def __init__(self, row, url):
        super().__init__()
        self.row = row
        self.url = url

    async def fetch_data(self, max_retries=10):
        for attempt in range(max_retries):
            try:
                async with aiohttp.ClientSession() as session:
                    response = await session.get(self.url)
                    response.raise_for_status()
                    response_text = await response.text()

                    # 正则表达式匹配逻辑
                    if '"Track shipment” link' in response_text:
                        Track_shipment = "Track shipment"
                    else:
                        Track_shipment = ""

                    modo = r'"productName"\s*:\s*"([^"]+)"'
                    time = r'"orderPlacedDate"\s*:\s*"([^"]+)"'
                    dd = r'/shop/order/detail/\d+/([^"/]+)'
                    status = r'"deliveryDate"\s*:\s*"([^"]+)"'
                    pattern = r'"currentStatus"\s*:\s*"([^"]+)"'

                    MODO = re.search(modo, response_text)
                    TIME = re.search(time, response_text)
                    DD = re.search(dd, response_text)
                    STATUS = re.search(status, response_text)
                    The_status_of_the_goods = re.search(pattern, response_text)

                    if MODO and TIME and DD and STATUS:
                        result_data = {
                            "productName": MODO.group(1),
                            "orderPlacedDate": TIME.group(1),
                            "orderNumber": DD.group(1),
                            "status": STATUS.group(1),
                            "Track_shipment": Track_shipment,
                            "The_status_of_the_goods": The_status_of_the_goods.group(1),
                        }
                        self.result.emit(self.row, result_data)
                        return  # 成功获取数据后退出函数

                    raise ValueError("Data not found in response")

            except Exception as e:
                if attempt < max_retries:
                    print(attempt)
                    await asyncio.sleep(2)
                    pass
                else:
                    self.result.emit(self.row, error_message)  # 发射错误信息
                error_message = {"error": f"Error: {e}"}
            finally:
                self.finished.emit()


class AsyncRequestWorker(QRunnable):
    def __init__(self, async_request):
        super().__init__()
        self.async_request = async_request

    def run(self):
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        loop.run_until_complete(self.async_request.fetch_data())


class InterFace(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.threadpool = QThreadPool()

    def initUI(self):
        self.setWindowTitle("Async HTTP Requests with PyQt6")
        self.setGeometry(100, 100, 1630, 600)

        h_layout_buttons = QHBoxLayout()
        self.button_import1 = QPushButton("导入", self)
        self.button_import2 = QPushButton("清理", self)
        self.button_import3 = QPushButton("查询", self)
        self.button_import4 = QPushButton("导出成功", self)
        self.button_import5 = QPushButton("导出失败", self)
        h_layout_buttons.addWidget(self.button_import1)
        h_layout_buttons.addWidget(self.button_import2)
        h_layout_buttons.addWidget(self.button_import3)
        h_layout_buttons.addWidget(self.button_import4)
        h_layout_buttons.addWidget(self.button_import5)
        h_layout_buttons.setAlignment(Qt.AlignmentFlag.AlignLeft)

        self.table_widget = QTableWidget(0, 7, self)
        self.table_widget.setHorizontalHeaderLabels(
            [
                "url",
                "modo",
                "order number",
                "order time",
                "status",
                "The_status_of_the_goods",
                "Track shipment",
            ]
        )
        self.table_widget.setShowGrid(True)
        self.table_widget.setColumnWidth(0, 500)
        self.table_widget.setColumnWidth(1, 300)
        self.table_widget.setColumnWidth(2, 100)
        self.table_widget.setColumnWidth(3, 100)
        self.table_widget.setColumnWidth(4, 200)
        self.table_widget.setColumnWidth(5, 200)
        self.table_widget.setColumnWidth(6, 200)

        v_layout = QVBoxLayout()
        v_layout.addLayout(h_layout_buttons)
        v_layout.addWidget(self.table_widget)
        self.setLayout(v_layout)

        self.button_import1.clicked.connect(self.Get_link)
        self.button_import2.clicked.connect(self.clear_table)
        self.button_import3.clicked.connect(self.on_query_clicked)
        self.button_import4.clicked.connect(self.export_data_to_xlsx_TRUE)
        self.button_import5.clicked.connect(self.export_data_to_xlsx_false)

    def on_query_clicked(self):
        for row in range(self.table_widget.rowCount()):
            for col in range(1, 7):
                self.table_widget.setItem(row, col, QTableWidgetItem(""))  # 清空列数据

        for row in range(self.table_widget.rowCount()):
            url_item = self.table_widget.item(row, 0)
            if url_item and url_item.text():
                async_request = AsyncRequest(row, url_item.text())
                async_request.result.connect(self.update_table)
                worker = AsyncRequestWorker(async_request)
                self.threadpool.start(worker)

    def update_table(self, row, result):
        if "error" in result:
            self.table_widget.setItem(row, 1, QTableWidgetItem(""))
            self.table_widget.setItem(row, 2, QTableWidgetItem(""))
            self.table_widget.setItem(row, 3, QTableWidgetItem(""))
            self.table_widget.setItem(row, 4, QTableWidgetItem(""))
            self.table_widget.setItem(row, 5, QTableWidgetItem(""))
            self.table_widget.setItem(row, 6, QTableWidgetItem(""))
        else:
            self.table_widget.setItem(
                row, 1, QTableWidgetItem(str(result["productName"]))
            )
            self.table_widget.setItem(
                row, 2, QTableWidgetItem(str(result["orderNumber"]))
            )
            self.table_widget.setItem(
                row, 3, QTableWidgetItem(str(result["orderPlacedDate"]))
            )
            self.table_widget.setItem(row, 4, QTableWidgetItem(str(result["status"])))
            self.table_widget.setItem(
                row, 5, QTableWidgetItem(str(result["The_status_of_the_goods"]))
            )
            self.table_widget.setItem(
                row, 6, QTableWidgetItem(str(result["Track_shipment"]))
            )

    def Get_link(self):
        clipboard = QApplication.clipboard().text()
        if clipboard:
            items = clipboard.split()
            self.table_widget.setRowCount(len(items))
            for i, item in enumerate(items):
                self.table_widget.setItem(i, 0, QTableWidgetItem(item))

    def clear_table(self):
        self.table_widget.clearContents()
        self.table_widget.setRowCount(0)

    def export_data_to_xlsx_TRUE(self):

        current_time = time.strftime("%Y%m%d_%H%M%S")
        file_name = f"成功_{current_time}.xlsx"
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        file_path = os.path.join(desktop_path, file_name)

        wb = Workbook()
        ws = wb.active

        headers = [
            "url",
            "modo",
            "order number",
            "order time",
            "status",
            "The status of the goods",
            "Track shipment",
        ]
        ws.append(headers)

        for row in range(self.table_widget.rowCount()):
            empty_columns = all(
                self.table_widget.item(row, col) is None
                or self.table_widget.item(row, col).text() == ""
                for col in [1, 2, 3, 4, 5, 6, 7]
            )
            if not empty_columns:
                row_data = [
                    (
                        self.table_widget.item(row, col).text()
                        if self.table_widget.item(row, col)
                        else ""
                    )
                    for col in range(self.table_widget.columnCount())
                ]
                ws.append(row_data)

        for col in ws.columns:
            max_length = max(
                len(str(cell.value)) for cell in col if cell.value is not None
            )
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col[0].column)
            ].width = max_length

        # 保存文件
        wb.save(file_path)

    def export_data_to_xlsx_false(self):
        current_time = time.strftime("%Y%m%d_%H%M%S")
        file_name = f"失败_{current_time}.xlsx"
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        file_path = os.path.join(desktop_path, file_name)

        wb = Workbook()
        ws = wb.active

        headers = ["url"]
        ws.append(headers)

        for row in range(self.table_widget.rowCount()):
            empty_columns = [
                self.table_widget.item(row, col) is None
                or self.table_widget.item(row, col).text() == ""
                for col in [1, 2, 3, 4]
            ]
            if any(empty_columns):  # 只要列表中有一个True值，即至少一列为空
                row_data = [
                    (
                        self.table_widget.item(row, col).text()
                        if self.table_widget.item(row, col)
                        else ""
                    )
                    for col in range(self.table_widget.columnCount())
                ]
                ws.append(row_data)

        for col in ws.columns:
            max_length = max(
                len(str(cell.value)) for cell in col if cell.value is not None
            )
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col[0].column)
            ].width = max_length

        wb.save(file_path)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = InterFace()
    window.show()
    sys.exit(app.exec())
