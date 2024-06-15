import pandas as pd
import re
import asyncio
from datetime import datetime
import os
from pathlib import Path
from two import APPLE
from concurrent.futures import ThreadPoolExecutor
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PySide6.QtCore import Qt, QRunnable, QThreadPool
from PySide6.QtWidgets import (
    QApplication,
    QMainWindow,
    QPushButton,
    QVBoxLayout,
    QWidget,
    QHBoxLayout,
    QTableWidget,
    QInputDialog,
    QTableWidgetItem,
    QMessageBox,
    QComboBox,
    QFileDialog,
)

# Pyinstaller -F -w Change_the_address.py


class Asynctask(QRunnable):
    def __init__(self, row, email, callback, **address_details):
        super().__init__()
        self.row = row
        self.email = email
        self.callback = callback
        self.address_details = address_details

    def run(self):
        loop = asyncio.new_event_loop()
        loop.run_until_complete(self.run_on_executor())
        loop.close()

    async def run_on_executor(self):
        loop = asyncio.get_running_loop()
        executor = ThreadPoolExecutor()
        apple = await loop.run_in_executor(executor, self.run_sync_method)
        self.callback(self.row, self.email, apple)

    def run_sync_method(self):
        apple = APPLE(**self.address_details)
        apple.t0()
        if hasattr(apple, "retry_status") and apple.retry_status == "失败":
            result = {"firstName": "失败"}
        else:
            billing_address_match = re.search(
                r'"billing-address":\s*{\s*"d":\s*{(.*?)}}',
                apple.return_10_status,
                re.DOTALL,
            ).group(1)

            def get_field_value(field_name):
                pattern = rf'"{field_name}":"([^"]*)"'
                match = re.search(pattern, billing_address_match)
                # return match.group(1) if match else None
                return match.group(1).rstrip(",") if match else None

            result = {
                "city": get_field_value("city"),
                "state": get_field_value("state"),
                "lastName": get_field_value("lastName"),
                "firstName": get_field_value("firstName"),
                "companyName": get_field_value("companyName"),
                "street": get_field_value("street"),
                "postalCode": get_field_value("postalCode"),
                "street2": get_field_value("street2"),
            }
            # print(result)
        return result


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Change_the_address")
        self.setGeometry(300, 300, 1150, 600)

        self.thread_pool = QThreadPool()

        main_layout = QVBoxLayout()
        button_layout = QHBoxLayout()
        form_layout = QHBoxLayout()
        table_layout = QHBoxLayout()
        main_layout.addLayout(button_layout)
        main_layout.addLayout(form_layout)
        main_layout.addLayout(table_layout)
        button_layout.setAlignment(Qt.AlignLeft)
        form_layout.setAlignment(Qt.AlignLeft)
        table_layout.setAlignment(Qt.AlignLeft)
        main_layout.setAlignment(Qt.AlignTop)

        self.tableWidget = QTableWidget()
        self.tableWidget.setColumnCount(11)
        table_layout.addWidget(self.tableWidget)
        self.tableWidget.setHorizontalHeaderLabels(
            [
                "账号",
                "密码",
                "链接",
                "城市",
                "州",
                "姓",
                "名",
                "公司名称",
                "街道",
                "邮政编码",
                "街道2",
            ]
        )

        self.button1 = QPushButton("导入")
        self.button2 = QPushButton("修改")
        self.button3 = QPushButton("导出")
        self.button4 = QPushButton("清空")
        self.button5 = QPushButton("导入地址")
        self.button6 = QPushButton("清空地址")
        button_layout.addWidget(self.button1)
        button_layout.addWidget(self.button2)
        button_layout.addWidget(self.button3)
        button_layout.addWidget(self.button4)
        button_layout.addWidget(self.button5)
        button_layout.addWidget(self.button6)
        self.button1.clicked.connect(self.on_click_button1)
        self.button2.clicked.connect(self.on_click_button2)
        self.button3.clicked.connect(self.on_click_button3)
        self.button4.clicked.connect(self.on_click_button4)
        self.button5.clicked.connect(self.on_click_button5)
        self.button6.clicked.connect(self.on_click_button6)

        self.comboBox1 = QComboBox()
        self.comboBox1.setFixedWidth(1000)
        form_layout.addWidget(self.comboBox1)
        self.comboBox1.setStyleSheet(
            """QComboBox {
            border: 1px solid lightgray;
            padding: 4px;
            font-size: 14pt;
            background-color: white; /* 确保有明确的背景色 */
        }

        QComboBox::drop-down {
            width: 15px;
            height: 15px;
        }

        QComboBox QAbstractItemView {
            background-color: white; /* 下拉列表背景色 */
            border: 1px solid lightgray; /* 下拉列表边框 */
        }

        QComboBox QAbstractItemView::item {
            padding: 4px 2px;
            min-height: 24px;
            border-bottom: 5px solid lightgray; /* 项之间的分隔线 */
        }

        QComboBox QAbstractItemView::item:selected { /* 选中的项 */
            background-color: lightblue; /* 选中项的背景色 */
            color: black; /* 选中项的文字颜色 */
        }

        QComboBox:on { /* 下拉框打开状态 */
            background-color: lightgray;
        }

        QComboBox::drop-down {
            subcontrol-origin: padding;
            subcontrol-position: right center;
        }
        """
        )

        widget = QWidget()
        widget.setLayout(main_layout)
        self.setCentralWidget(widget)

    def on_click_button1(self):
        text, ok = QInputDialog.getMultiLineText(
            self, "导入账号,密码,超链接", "请输入账号,密码,超链接(每行一个，用空格分隔)"
        )
        if ok and text:
            lines = text.split("\n")
            for line in lines:
                line = line.strip()
                if line:
                    email, password, url = line.split()
                    self.add_email_password_and_url_to_table(email, password, url)

    def add_email_password_and_url_to_table(self, email, password, url):
        row_position = self.tableWidget.rowCount()
        self.tableWidget.insertRow(row_position)
        email_item = QTableWidgetItem(email)
        password_item = QTableWidgetItem(password)
        url_item = QTableWidgetItem(url)
        self.tableWidget.setItem(row_position, 0, email_item)
        self.tableWidget.setItem(row_position, 1, password_item)
        self.tableWidget.setItem(row_position, 2, url_item)

    # def on_click_button2(self):
    #     for row in range(self.tableWidget.rowCount()):
    #         if self.tableWidget.item(row, 0) and self.tableWidget.item(row, 1):
    #             email = self.tableWidget.item(row, 0).text()
    #             selected_text = self.comboBox1.currentText()
    #             parts = selected_text.split('----')
    #             params = {
    #                 "name":self.tableWidget.item(row, 0).text(),
    #                 "pwd":self.tableWidget.item(row, 1).text(),
    #                 "url": self.tableWidget.item(row, 2).text(),
    #                 "city": parts[2].strip(),#城市
    #                 "state": parts[3].strip(),#州
    #                 "lastName": parts[0].strip(),#姓
    #                 "firstName": parts[1].strip(),#名
    #                 "companyName": '',#"公司名称"
    #                 "street": parts[4].strip(),#"街道"
    #                 "postalCode": parts[6].strip(),#"邮政编码"
    #                 "street2": parts[5].strip(),#"街道2"
    #             }
    #             # params = {**params}
    #             task = Asynctask(row, email, self.update_table_item, **params)
    #             self.thread_pool.start(task)

    def on_click_button2(self):
        # 检查comboBox1是否为空
        selected_text = self.comboBox1.currentText()
        if not selected_text:
            QMessageBox.warning(self, "警告", "没有导入地址，请先导入地址再执行修改。")
            return  # 如果comboBox为空，则不执行下面的代码

        # 标志变量，用于控制循环是否继续
        continue_loop = True
        for row in range(self.tableWidget.rowCount()):
            # 检查每一行的第一列、第二列和第三列是否有内容
            if not (
                self.tableWidget.item(row, 0)
                and self.tableWidget.item(row, 1)
                and self.tableWidget.item(row, 2)
            ):
                QMessageBox.warning(
                    self, "警告", "请确保表格的每一行都有账号、密码和链接信息。"
                )
                continue_loop = False  # 设置标志变量，退出循环
                break  # 退出当前循环

            if continue_loop:  # 只有当continue_loop为True时，才继续执行
                email = self.tableWidget.item(row, 0).text()
                parts = selected_text.split("----")
                params = {
                    "name": self.tableWidget.item(row, 0).text(),
                    "pwd": self.tableWidget.item(row, 1).text(),
                    "url": self.tableWidget.item(
                        row, 2
                    ).text(),  # 注意：这里应该是索引2，而不是'url'
                    "city": parts[2].strip(),  # 城市
                    "state": parts[3].strip(),  # 州
                    "lastName": parts[0].strip(),  # 姓
                    "firstName": parts[1].strip(),  # 名
                    "companyName": "",  # "公司名称"
                    "street": parts[4].strip(),  # "街道"
                    "postalCode": parts[6].strip(),  # "邮政编码"
                    "street2": parts[5].strip(),  # "街道2"
                }
                task = Asynctask(row, email, self.update_table_item, **params)
                self.thread_pool.start(task)

    def update_table_item(self, row, email, result):
        if result.get("firstName") == "失败":
            self.tableWidget.setItem(row, 3, QTableWidgetItem(result["firstName"]))
        else:
            column_mapping = {
                "city": 3,
                "state": 4,
                "lastName": 5,
                "firstName": 6,
                "companyName": 7,
                "street": 8,
                "postalCode": 9,
                "street2": 10,
            }
            for key, col_index in column_mapping.items():
                self.tableWidget.setItem(
                    row, col_index, QTableWidgetItem(result.get(key, ""))
                )

    def on_click_button3(self):
        desktop_path = str(Path.home() / "Desktop")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"地址_{timestamp}.xlsx"
        file_path = os.path.join(desktop_path, filename)
        df = self.tableWidgetToDataFrame(self.tableWidget)
        df.to_excel(file_path, index=False, engine="openpyxl")
        wb = load_workbook(file_path)
        ws = wb.active
        for col in range(1, ws.max_column + 1):
            max_length = 0
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col)].width = max_length + 5
        wb.save(file_path)
        wb.close()
        QMessageBox.information(self, "导出完成", f"数据已导出到 {file_path}")

    def tableWidgetToDataFrame(self, tableWidget):
        num_rows = tableWidget.rowCount()
        num_cols = tableWidget.columnCount()
        columns = [tableWidget.horizontalHeaderItem(i).text() for i in range(num_cols)]
        df = pd.DataFrame(columns=columns)
        for i in range(num_rows):
            row_data = [
                tableWidget.item(i, j).text() if tableWidget.item(i, j) else ""
                for j in range(num_cols)
            ]
            df.loc[i] = row_data
        return df

    def on_click_button4(self):
        self.tableWidget.clearContents()
        self.tableWidget.setRowCount(0)

    def on_click_button5(self):
        try:
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "选择文件",
                QApplication.instance().applicationDirPath(),
                "Excel Files (*.xlsx);;All Files (*)",
            )
            if file_path:
                try:
                    dtypes = {"邮编": str}
                    df = pd.read_excel(file_path, engine="openpyxl", dtype=dtypes)
                    self.comboBox1.clear()
                    df.fillna("     ", inplace=True)
                    for index, row in df.iterrows():
                        full_address = (
                            f"{row['姓']}----"
                            f"{row['名']}----"
                            f"{row['城市']}----"
                            f"{row['地区']}----"
                            f"{row['地址一']}----"
                            f"{row['地址二']}----"
                            f"{row['邮编']}"
                        )
                        self.comboBox1.addItem(full_address.strip())
                except Exception as e:
                    print(f"读取文件时发生错误: {e}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"读取文件时发生错误: {e}")

    def on_click_button6(self):
        self.comboBox1.clear()


def main():
    app = QApplication([])
    window = MainWindow()
    window.show()
    app.exec()


if __name__ == "__main__":
    main()
