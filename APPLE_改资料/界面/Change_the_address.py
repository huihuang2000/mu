import pandas as pd
import asyncio
from datetime import datetime
import os
import re
from pathlib import Path
from one import APPLE
from concurrent.futures import ThreadPoolExecutor
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PySide6.QtCore import Qt,QRunnable, QThreadPool,QStandardPaths
from PySide6.QtWidgets import (
    QApplication,
    QMainWindow,
    QPushButton,
    QVBoxLayout,
    QWidget,
    QLineEdit,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QTableWidget,
    QInputDialog,
    QTableWidgetItem,
    QMessageBox,
    QComboBox,
    QFileDialog
)


class Asynctask(QRunnable):
    def __init__(self, row, email, callback, **address_details):
        super().__init__()
        self.row = row
        self.email = email
        self.callback = callback
        self.address_details = address_details

    def run(self):
        loop = asyncio.new_event_loop()
        loop.run_until_complete(self.run_on_executor())
        loop.close()

    async def run_on_executor(self):
        loop = asyncio.get_running_loop()
        executor = ThreadPoolExecutor()
        apple = await loop.run_in_executor(executor, self.run_sync_method)
        self.callback(self.row, self.email, apple)

    def run_sync_method(self):
        apple = APPLE(**self.address_details)
        apple.t0()
        if hasattr(apple, "stast") and apple.stast == "账号异常，无法获取必要的信息。":
            result = {"firstName": "账号异常"}
        else:
            result = {
                "firstName": apple.firstName,
                "lastName": apple.lastName,
                "companyName": apple.companyName,
                "street": apple.street,
                "street2": apple.street2,
                "postalCode": apple.postalCode,
                "city": apple.city,
                "state": apple.state,
                "countryCode": apple.countryCode,
                "fullDaytimePhone": apple.fullDaytimePhone,
            }
        return result


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Change_the_address")
        self.setGeometry(300, 300, 1200, 600)

        self.thread_pool = QThreadPool()

        main_layout = QVBoxLayout()
        button_layout = QHBoxLayout()
        label_layout = QHBoxLayout()
        form_layout = QHBoxLayout()
        table_layout = QHBoxLayout()
        main_layout.addLayout(button_layout)
        main_layout.addLayout(label_layout)
        main_layout.addLayout(form_layout)
        main_layout.addLayout(table_layout)
        button_layout.setAlignment(Qt.AlignLeft)
        label_layout.setAlignment(Qt.AlignLeft)
        form_layout.setAlignment(Qt.AlignLeft)
        table_layout.setAlignment(Qt.AlignLeft)
        main_layout.setAlignment(Qt.AlignTop)

        self.tableWidget = QTableWidget()
        self.tableWidget.setColumnCount(12)
        # self.tableWidget.setFixedWidth(1200)
        # self.tableWidget.setFixedHeight(500)
        table_layout.addWidget(self.tableWidget)
        self.tableWidget.setHorizontalHeaderLabels(
            [
                "邮箱",
                "密码",
                "名",
                "姓",
                "公司名称",
                "街道地址",
                "公寓",
                "邮政编码",
                "城市",
                "州",
                "国家地区",
                "电话号码",
            ]
        )

        self.button1 = QPushButton("导入")
        self.button2 = QPushButton("修改")
        self.button3 = QPushButton("导出")
        self.button4 = QPushButton("清空")
        self.button5 = QPushButton("导入地址")
        self.button6 = QPushButton("清空地址")
        button_layout.addWidget(self.button1)
        button_layout.addWidget(self.button2)
        button_layout.addWidget(self.button3)
        button_layout.addWidget(self.button4)
        button_layout.addWidget(self.button5)
        button_layout.addWidget(self.button6)
        self.button1.clicked.connect(self.on_click_button1)
        self.button2.clicked.connect(self.on_click_button2)
        self.button3.clicked.connect(self.on_click_button3)
        self.button4.clicked.connect(self.on_click_button4)
        self.button5.clicked.connect(self.on_click_button5)
        self.button6.clicked.connect(self.on_click_button6)

        self.label1 = QLabel("名")
        self.label1.setFixedWidth(100)
        label_layout.addWidget(self.label1)
        self.comboBox1 = QComboBox()
        self.comboBox1.setFixedWidth(1000)
        form_layout.addWidget(self.comboBox1)


        self.label2 = QLabel("姓")
        self.label2.setFixedWidth(100)
        label_layout.addWidget(self.label2)
        # self.comboBox2 = QComboBox()
        # self.comboBox2.setFixedWidth(100)
        # self.comboBox2.addItem("选项1", "值1")
        # self.comboBox2.addItem("选项2", "值2")
        # form_layout.addWidget(self.comboBox2)

        self.label3 = QLabel("公司名称")
        self.label3.setFixedWidth(100)
        label_layout.addWidget(self.label3)
        # self.comboBox3 = QComboBox()
        # self.comboBox3.setFixedWidth(100)
        # self.comboBox3.addItem("选项1", "值1")
        # self.comboBox3.addItem("选项2", "值2")
        # form_layout.addWidget(self.comboBox3)

        self.label4 = QLabel("街道地址")
        self.label4.setFixedWidth(100)
        label_layout.addWidget(self.label4)
        # self.comboBox4 = QComboBox()
        # self.comboBox4.setFixedWidth(100)
        # self.comboBox4.addItem("选项1", "值1")
        # self.comboBox4.addItem("选项2", "值2")
        # form_layout.addWidget(self.comboBox4)

        self.label5 = QLabel("公寓")
        self.label5.setFixedWidth(100)
        label_layout.addWidget(self.label5)
        # self.comboBox5 = QComboBox()
        # self.comboBox5.setFixedWidth(100)
        # self.comboBox5.addItem("选项1", "值1")
        # self.comboBox5.addItem("选项2", "值2")
        # form_layout.addWidget(self.comboBox5)

        self.label6 = QLabel("邮政编码")
        self.label6.setFixedWidth(100)
        label_layout.addWidget(self.label6)
        # self.comboBox6 = QComboBox()
        # self.comboBox6.setFixedWidth(100)
        # self.comboBox6.addItem("选项1", "值1")
        # self.comboBox6.addItem("选项2", "值2")
        # form_layout.addWidget(self.comboBox6)

        self.label7 = QLabel("城市")
        self.label7.setFixedWidth(100)
        label_layout.addWidget(self.label7)
        # self.comboBox7 = QComboBox()
        # self.comboBox7.setFixedWidth(100)
        # self.comboBox7.addItem("选项1", "值1")
        # self.comboBox7.addItem("选项2", "值2")
        # form_layout.addWidget(self.comboBox7)


        self.label8 = QLabel("州")
        self.label8.setFixedWidth(100)
        label_layout.addWidget(self.label8)
        # self.comboBox8 = QComboBox()
        # self.comboBox8.setFixedWidth(100)
        # self.comboBox8.addItem("选项1", "值1")
        # self.comboBox8.addItem("选项2", "值2")
        # form_layout.addWidget(self.comboBox8)

        self.label9 = QLabel("国家地区")
        self.label9.setFixedWidth(100)
        label_layout.addWidget(self.label9)
        # self.comboBox9 = QComboBox()
        # self.comboBox9.setFixedWidth(100)
        # self.comboBox9.addItem("选项1", "值1")
        # self.comboBox9.addItem("选项2", "值2")
        # form_layout.addWidget(self.comboBox9)

        self.label10 = QLabel("电话号码")
        self.label10.setFixedWidth(100)
        label_layout.addWidget(self.label10)
        # self.comboBox10 = QComboBox()
        # self.comboBox10.setFixedWidth(100)
        # self.comboBox10.addItem("选项1", "值1")
        # self.comboBox10.addItem("选项2", "值2")
        # form_layout.addWidget(self.comboBox10)

        widget = QWidget()
        widget.setLayout(main_layout)
        self.setCentralWidget(widget)

    def on_click_button1(self):
        text, ok = QInputDialog.getMultiLineText(
            self, "导入账号密码", "请输入邮箱和密码(每行一个，用空格分隔)"
        )
        if ok and text:
            lines = text.split("\n")
            for line in lines:
                line = line.strip()
                if line:
                    email, password, *rest = line.split(maxsplit=1)
                    if rest:
                        print(f"警告：格式错误 - 多余的空格被忽略。行内容：{line}")
                    self.add_email_and_password_to_table(email, password)

    def add_email_and_password_to_table(self, email, password):
        row_position = self.tableWidget.rowCount()
        self.tableWidget.insertRow(row_position)
        email_item = QTableWidgetItem(email)
        password_item = QTableWidgetItem(password)
        self.tableWidget.setItem(row_position, 0, email_item)
        self.tableWidget.setItem(row_position, 1, password_item)

    def on_click_button2(self):
        for row in range(self.tableWidget.rowCount()):
            if self.tableWidget.item(row, 0) and self.tableWidget.item(row, 1):
                email = self.tableWidget.item(row, 0).text()
                password = self.tableWidget.item(row, 1).text()
                address_details = {
                    "name":email,
                    "pwd":password,
                    "fullDaytimePhone": self.comboBox1.currentText(),
                    "street2": self.comboBox2.currentText(),
                    "lastName": self.comboBox3.currentText(),
                    "firstName": self.comboBox4.currentText(),
                    "companyName": self.comboBox5.currentText(),
                    "street": self.comboBox6.currentText(),
                    "city": self.comboBox7.currentText(),
                    "state": self.comboBox8.currentText(),
                    "postalCode": self.comboBox9.currentText(),
                    "countryCode": self.comboBox10.currentText(),
                }
                params = {
                    "name": email,
                    "pwd": password,
                    "fullDaytimePhone": "111111-2222",
                    "street2": "212231312",
                    "lastName": "lllll3333",
                    "firstName": "hhhjjj11166666",
                    "companyName": "122332",
                    "street": "777776667777",
                    "city": "Albany",
                    "state": "CH",
                    "postalCode": "11111-1111",
                    "countryCode": "CH",
                }
                params = {**params}
                task = Asynctask(row, email, self.update_table_item, **params)
                self.thread_pool.start(task)

    def update_table_item(self, row, email, result):
        if result.get("firstName") == "账号异常":
            self.tableWidget.setItem(row, 2, QTableWidgetItem(result["firstName"]))
        else:
            if self.tableWidget.item(row, 0).text() == email:
                self.tableWidget.setItem(
                    row, 2, QTableWidgetItem(result.get("firstName", ""))
                )
                self.tableWidget.setItem(
                    row, 3, QTableWidgetItem(result.get("lastName", ""))
                )
                self.tableWidget.setItem(
                    row, 4, QTableWidgetItem(result.get("companyName", ""))
                )
                self.tableWidget.setItem(
                    row, 5, QTableWidgetItem(result.get("street", ""))
                )
                self.tableWidget.setItem(
                    row, 6, QTableWidgetItem(result.get("street2", ""))
                )
                self.tableWidget.setItem(
                    row, 7, QTableWidgetItem(result.get("postalCode", ""))
                )
                self.tableWidget.setItem(
                    row, 8, QTableWidgetItem(result.get("city", ""))
                )
                self.tableWidget.setItem(
                    row, 9, QTableWidgetItem(result.get("state", ""))
                )
                self.tableWidget.setItem(
                    row, 10, QTableWidgetItem(result.get("countryCode", ""))
                )
                self.tableWidget.setItem(
                    row, 11, QTableWidgetItem(result.get("fullDaytimePhone", ""))
                )

    def on_click_button3(self):
        desktop_path = str(Path.home() / "Desktop")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"地址_{timestamp}.xlsx"
        file_path = os.path.join(desktop_path, filename)
        df = self.tableWidgetToDataFrame(self.tableWidget)
        df.to_excel(file_path, index=False, engine="openpyxl")
        wb = load_workbook(file_path)
        ws = wb.active
        for col in range(1, ws.max_column + 1):
            max_length = 0
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col)].width = max_length + 5
        wb.save(file_path)
        wb.close()
        QMessageBox.information(self, "导出完成", f"数据已导出到 {file_path}")

    def tableWidgetToDataFrame(self, tableWidget):
        num_rows = tableWidget.rowCount()
        num_cols = tableWidget.columnCount()
        columns = [tableWidget.horizontalHeaderItem(i).text() for i in range(num_cols)]
        df = pd.DataFrame(columns=columns)
        for i in range(num_rows):
            row_data = [
                tableWidget.item(i, j).text() if tableWidget.item(i, j) else ""
                for j in range(num_cols)
            ]
            df.loc[i] = row_data
        return df

    def on_click_button4(self):
        self.tableWidget.clearContents()
        self.tableWidget.setRowCount(0)


    def on_click_button5(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "选择文件",
            QApplication.instance().applicationDirPath(),
            "Excel Files (*.xlsx);;All Files (*)"
        )
        if file_path:
            try:
                df = pd.read_excel(file_path, engine='openpyxl')
                self.comboBox1.clear()
                df.fillna('     ', inplace=True)
                for index, row in df.iterrows():
                    full_address = (
                        f"{row['姓']}----"  # 姓
                        f"{row['名']}----"  # 名
                        f"{row['城市']}----"  # 城市
                        f"{row['地区']}----"  # 地区
                        f"{row['地址一']}----"  # 地址一
                        f"{row['地址二']}----"  # 地址二
                        f"{row['邮编']}"  # 邮编
                    )
                    self.comboBox1.addItem(full_address.strip())
            except Exception as e:
                print(f"读取文件时发生错误: {e}")

    def on_click_button6(self):
        self.comboBox1.clear()


def main():
    app = QApplication([])
    window = MainWindow()
    window.show()
    app.exec()


if __name__ == "__main__":
    main()
